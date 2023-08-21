import openpyxl

source_workbook = openpyxl.load_workbook('source workbook')
destination_workbook = openpyxl.load_workbook('destination workbook')

source_sheet = source_workbook['Sheet1'] 
destination_sheet = destination_workbook['Sheet2']  

for source_row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1, max_col=source_sheet.max_column):

    source_value1 = source_row[0].value
    
    for dest_row in destination_sheet.iter_rows(min_row=1, max_row=destination_sheet.max_row, min_col=1, max_col=destination_sheet.max_column):

        dest_value1 = dest_row[0].value

        if dest_value1 == source_value1 and dest_value1 is not None:

            for i in range(len(dest_row)):
                # print(dest_value1)

                destination_sheet.cell(row=dest_row[0].value+1, column=i+1, value=source_row[i].value)

destination_workbook.save('test.xlsx')
